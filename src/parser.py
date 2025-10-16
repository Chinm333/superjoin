"""
Spreadsheet Content Parser

This module handles reading and parsing spreadsheet files (Excel/Google Sheets)
to extract formulas, values, headers, and cell relationships.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any
import re
from dataclasses import dataclass
from pathlib import Path


@dataclass
class CellInfo:
    """Information about a single cell"""
    sheet_name: str
    cell_address: str
    value: Any
    formula: Optional[str] = None
    data_type: str = "unknown"
    is_header: bool = False
    is_formula: bool = False
    row: int = 0
    col: int = 0


@dataclass
class SheetInfo:
    """Information about a spreadsheet sheet"""
    name: str
    cells: List[CellInfo]
    headers: List[str]
    data_range: Tuple[int, int, int, int]  # (min_row, max_row, min_col, max_col)
    formulas: List[CellInfo]
    values: List[CellInfo]


class SpreadsheetParser:
    """Parser for Excel and Google Sheets files"""
    
    def __init__(self):
        self.workbook = None
        self.sheets_info: Dict[str, SheetInfo] = {}
    
    def load_excel_file(self, file_path: str) -> Dict[str, SheetInfo]:
        """
        Load and parse an Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary mapping sheet names to SheetInfo objects
        """
        try:
            self.workbook = load_workbook(file_path, data_only=False)
            self.sheets_info = {}
            
            for sheet_name in self.workbook.sheetnames:
                sheet_info = self._parse_sheet(sheet_name)
                self.sheets_info[sheet_name] = sheet_info
                
            return self.sheets_info
            
        except Exception as e:
            raise Exception(f"Error loading Excel file {file_path}: {str(e)}")
    
    def load_google_sheets(self, sheet_id: str, credentials_path: Optional[str] = None) -> Dict[str, SheetInfo]:
        """
        Load and parse a Google Sheets file
        
        Args:
            sheet_id: Google Sheets ID from URL
            credentials_path: Path to service account credentials (optional)
            
        Returns:
            Dictionary mapping sheet names to SheetInfo objects
        """
        # TODO: Implement Google Sheets integration
        # This would require google-auth and google-api-python-client
        raise NotImplementedError("Google Sheets integration not yet implemented")
    
    def _parse_sheet(self, sheet_name: str) -> SheetInfo:
        """Parse a single sheet and extract all relevant information"""
        sheet = self.workbook[sheet_name]
        
        cells = []
        headers = []
        formulas = []
        values = []
        
        # Get the used range
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
            # Empty sheet
            return SheetInfo(
                name=sheet_name,
                cells=[],
                headers=[],
                data_range=(1, 1, 1, 1),
                formulas=[],
                values=[]
            )
        
        min_row, max_row = 1, sheet.max_row
        min_col, max_col = 1, sheet.max_column
        
        # Parse all cells
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row, col)
                
                if cell.value is not None:
                    cell_info = self._create_cell_info(sheet_name, cell, row, col)
                    cells.append(cell_info)
                    
                    if cell_info.is_formula:
                        formulas.append(cell_info)
                    else:
                        values.append(cell_info)
                    
                    # Identify headers (usually in first few rows)
                    if row <= 3 and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                        if self._is_likely_header(cell.value, row, col, sheet):
                            cell_info.is_header = True
                            headers.append(cell.value.strip())
        
        return SheetInfo(
            name=sheet_name,
            cells=cells,
            headers=headers,
            data_range=(min_row, max_row, min_col, max_col),
            formulas=formulas,
            values=values
        )
    
    def _create_cell_info(self, sheet_name: str, cell, row: int, col: int) -> CellInfo:
        """Create a CellInfo object from an openpyxl cell"""
        cell_address = f"{openpyxl.utils.get_column_letter(col)}{row}"
        
        # Determine if it's a formula
        is_formula = cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('='))
        formula = cell.value if is_formula else None
        
        # Determine data type
        if is_formula:
            data_type = "formula"
        elif isinstance(cell.value, (int, float)):
            data_type = "number"
        elif isinstance(cell.value, str):
            data_type = "text"
        elif isinstance(cell.value, bool):
            data_type = "boolean"
        elif cell.value is None:
            data_type = "empty"
        else:
            data_type = "unknown"
        
        return CellInfo(
            sheet_name=sheet_name,
            cell_address=cell_address,
            value=cell.value,
            formula=formula,
            data_type=data_type,
            is_formula=is_formula,
            row=row,
            col=col
        )
    
    def _is_likely_header(self, value: str, row: int, col: int, sheet) -> bool:
        """Determine if a cell is likely a header"""
        if not isinstance(value, str):
            return False
        
        value = value.strip()
        
        # Common header patterns
        header_patterns = [
            r'^[A-Z][a-z]+(\s+[A-Z][a-z]+)*$',  # Title Case
            r'^[A-Z_]+$',  # ALL_CAPS
            r'.*%$',  # Percentage headers
            r'.*\(.*\)$',  # Headers with units in parentheses
        ]
        
        for pattern in header_patterns:
            if re.match(pattern, value):
                return True
        
        # Check if it's in the first row or column
        if row == 1 or col == 1:
            return True
        
        # Check if surrounding cells are also text (indicating header row)
        if row <= 2:
            surrounding_text_count = 0
            for r in range(max(1, row-1), min(sheet.max_row+1, row+2)):
                for c in range(max(1, col-1), min(sheet.max_column+1, col+2)):
                    if r != row or c != col:
                        cell_val = sheet.cell(r, c).value
                        if isinstance(cell_val, str) and len(cell_val.strip()) > 0:
                            surrounding_text_count += 1
            
            if surrounding_text_count >= 2:
                return True
        
        return False
    
    def get_formula_dependencies(self, formula: str) -> List[str]:
        """Extract cell references from a formula"""
        if not formula or not formula.startswith('='):
            return []
        
        # Pattern to match cell references like A1, B2:C5, Sheet1!A1, etc.
        pattern = r'[A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?'
        matches = re.findall(pattern, formula)
        
        # Also match sheet references like Sheet1!A1
        sheet_pattern = r'[A-Za-z0-9_]+![A-Za-z]+[0-9]+'
        sheet_matches = re.findall(sheet_pattern, formula)
        
        return list(set(matches + sheet_matches))
    
    def get_sheet_summary(self, sheet_name: str) -> Dict[str, Any]:
        """Get a summary of a sheet's content"""
        if sheet_name not in self.sheets_info:
            return {}
        
        sheet_info = self.sheets_info[sheet_name]
        
        return {
            "name": sheet_name,
            "total_cells": len(sheet_info.cells),
            "formula_count": len(sheet_info.formulas),
            "header_count": len(sheet_info.headers),
            "data_range": sheet_info.data_range,
            "headers": sheet_info.headers[:10],  # First 10 headers
            "formula_types": self._analyze_formula_types(sheet_info.formulas)
        }
    
    def _analyze_formula_types(self, formulas: List[CellInfo]) -> Dict[str, int]:
        """Analyze the types of formulas in a sheet"""
        formula_types = {}
        
        for cell_info in formulas:
            if cell_info.formula:
                formula = cell_info.formula.upper()
                
                # Categorize formula types
                if any(func in formula for func in ['SUM', 'SUMIF', 'SUMIFS']):
                    formula_types['SUM'] = formula_types.get('SUM', 0) + 1
                elif any(func in formula for func in ['AVERAGE', 'AVERAGEIF', 'AVERAGEIFS']):
                    formula_types['AVERAGE'] = formula_types.get('AVERAGE', 0) + 1
                elif any(func in formula for func in ['COUNT', 'COUNTIF', 'COUNTIFS']):
                    formula_types['COUNT'] = formula_types.get('COUNT', 0) + 1
                elif any(func in formula for func in ['IF', 'IFS']):
                    formula_types['IF'] = formula_types.get('IF', 0) + 1
                elif any(func in formula for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
                    formula_types['LOOKUP'] = formula_types.get('LOOKUP', 0) + 1
                elif '/' in formula and '*' in formula:
                    formula_types['CALCULATION'] = formula_types.get('CALCULATION', 0) + 1
                else:
                    formula_types['OTHER'] = formula_types.get('OTHER', 0) + 1
        
        return formula_types


if __name__ == "__main__":
    parser = SpreadsheetParser()
    try:
        print("SpreadsheetParser initialized successfully!")
        print("Ready to parse Excel files when provided.")
        
    except Exception as e:
        print(f"Error: {e}")

