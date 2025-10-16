"""
Non-Interactive Demo for Semantic Spreadsheet Search Engine

This version runs all demo queries automatically without user input.
"""

import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from main import SemanticSpreadsheetSearch
from ranking import ResultFormat


def create_sample_data():
    """Create a sample Excel file for demonstration"""
    try:
        import pandas as pd
        import openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("Financial Summary")
        headers = ["Metric", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024", "Total"]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
            
        data = [
            ["Revenue", 100000, 120000, 110000, 130000, "=SUM(B2:E2)"],
            ["Cost of Goods Sold", 60000, 72000, 66000, 78000, "=SUM(B3:E3)"],
            ["Gross Profit", "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3", "=SUM(B4:E4)"],
            ["Gross Profit Margin %", "=B4/B2*100", "=C4/C2*100", "=D4/D2*100", "=E4/E2*100", "=F4/F2*100"],
            ["Operating Expenses", 25000, 28000, 26000, 30000, "=SUM(B6:E6)"],
            ["Operating Profit", "=B4-B6", "=C4-C6", "=D4-D6", "=E4-E6", "=SUM(B7:E7)"],
            ["Operating Margin %", "=B7/B2*100", "=C7/C2*100", "=D7/D2*100", "=E7/E2*100", "=F7/F2*100"]
        ]
        
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws1.cell(row=row, column=col, value=value)

        ws2 = wb.create_sheet("Budget vs Actual")
        headers2 = ["Metric", "Budget", "Actual", "Variance", "Variance %"]
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=header)
        
        data2 = [
            ["Revenue", 450000, 460000, "=C2-B2", "=D2/B2*100"],
            ["COGS", 270000, 276000, "=C3-B3", "=D3/B3*100"],
            ["Gross Profit", "=B2-B3", "=C2-C3", "=C4-B4", "=D4/B4*100"],
            ["Operating Expenses", 100000, 95000, "=C5-B5", "=D5/B5*100"],
            ["Net Profit", "=B4-B5", "=C4-C5", "=C6-B6", "=D6/B6*100"]
        ]
        
        for row, row_data in enumerate(data2, 2):
            for col, value in enumerate(row_data, 1):
                ws2.cell(row=row, column=col, value=value)
        
        ws3 = wb.create_sheet("Growth Analysis")
        
        headers3 = ["Period", "Revenue", "YoY Growth %", "QoQ Growth %"]
        for col, header in enumerate(headers3, 1):
            ws3.cell(row=1, column=col, value=header)

        data3 = [
            ["Q1 2023", 90000, "", ""],
            ["Q2 2023", 100000, "=(B3-B2)/B2*100", ""],
            ["Q3 2023", 95000, "=(B4-B3)/B3*100", "=(B4-B3)/B3*100"],
            ["Q4 2023", 110000, "=(B5-B4)/B4*100", "=(B5-B4)/B4*100"],
            ["Q1 2024", 100000, "=(B6-B2)/B2*100", "=(B6-B5)/B5*100"],
            ["Q2 2024", 120000, "=(B7-B3)/B3*100", "=(B7-B6)/B6*100"]
        ]
        
        for row, row_data in enumerate(data3, 2):
            for col, value in enumerate(row_data, 1):
                ws3.cell(row=row, column=col, value=value)
        
        sample_file = "sample_financial_data.xlsx"
        wb.save(sample_file)
        print(f"Created sample file: {sample_file}")
        return sample_file
        
    except ImportError:
        print("Error: openpyxl not installed. Please install it with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"Error creating sample file: {e}")
        return None


def run_demo():
    """Run the semantic search demo"""
    print("=" * 60)
    print("SEMANTIC SPREADSHEET SEARCH ENGINE - DEMO")
    print("=" * 60)
    
    search_engine = SemanticSpreadsheetSearch()
    
    print("\n1. Creating sample financial data...")
    sample_file = create_sample_data()
    
    if not sample_file:
        print("Could not create sample file. Please provide your own Excel file.")
        return
    
    print(f"\n2. Loading sample file: {sample_file}")
    load_result = search_engine.load_spreadsheet(sample_file)
    
    if not load_result["success"]:
        print(f"Error loading file: {load_result['error']}")
        return
    
    print(f"+ Loaded {load_result['total_sheets']} sheets:")
    for sheet_name, summary in load_result["summaries"].items():
        print(f"  - {sheet_name}: {summary['total_cells']} cells, {summary['formula_count']} formulas")

    print(f"\n3. Business concepts found:")
    concepts = search_engine.get_concept_overview()
    for concept, data in concepts.items():
        print(f"  - {concept}: {data['total_occurrences']} occurrences across {len(data['sheets'])} sheets")

    demo_queries = [
        "find all revenue calculations",
        "show profitability metrics",
        "where are my margin analyses?",
        "find budget vs actual comparisons",
        "show growth rate calculations",
        "find all percentage formulas"
    ]
    
    print(f"\n4. Running demo queries:")
    print("-" * 40)
    
    for i, query in enumerate(demo_queries, 1):
        print(f"\nQuery {i}: '{query}'")
        print("-" * 30)
        
        results = search_engine.search(query, max_results=3)
        print(results)
        
        if i < len(demo_queries):
            print("\n" + "="*50)
    
    print(f"\n5. Testing different output formats:")
    print("=" * 50)
    
    test_query = "find margin calculations"
    formats = [
        (ResultFormat.HUMAN_READABLE, "Human Readable"),
        (ResultFormat.GROUPED, "Grouped by Concept"),
        (ResultFormat.STRUCTURED, "Structured JSON")
    ]
    
    for format_type, format_name in formats:
        print(f"\n{format_name} Format:")
        print("-" * 20)
        results = search_engine.search(test_query, format_type=format_type, max_results=2)
        print(results[:500] + "..." if len(results) > 500 else results)

    try:
        os.remove(sample_file)
        print(f"\nCleaned up sample file: {sample_file}")
    except:
        pass
    
    print("\n" + "=" * 60)
    print("DEMO COMPLETED SUCCESSFULLY!")
    print("=" * 60)


if __name__ == "__main__":
    run_demo()

